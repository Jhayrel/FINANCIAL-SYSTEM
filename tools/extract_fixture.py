#!/usr/bin/env python3
"""
Extract a test fixture from the original Excel workbook.

READ-ONLY. This script never opens the workbook for writing and never moves,
renames or deletes anything under "MY THINGS/". See ../CLAUDE.md.

It produces app/src/fixtures/excel-fixture.json containing:
  - transactions      the ledger, money as integer centavos
  - deleted           the soft-delete recycle bin
  - budgets           per-year, per-month, two tracks
  - reference         wallets / savings / bills / subs / revenue / spending types
  - expected          figures read straight out of the workbook's own cells,
                      so parity tests assert against what Excel actually
                      displayed rather than against anything hand-transcribed

SECURITY: CATEGORIES column U holds live API keys. This script never reads
that column. Do not add it.

Usage:  python tools/extract_fixture.py
"""

from __future__ import annotations

import json
import sys
import warnings
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

warnings.filterwarnings("ignore")  # workbook uses unsupported DV extensions

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = (
    ROOT
    / "MY THINGS"
    / "ORIGINAL EXCEL WITH VBA"
    / "COPY - 2026 FINANCIAL AND BUDGETING SYSTEM.xlsm"
)
OUT = ROOT / "app" / "src" / "fixtures" / "excel-fixture.json"

DATA_START_ROW = 7
MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


# ── helpers ────────────────────────────────────────────────────────────────

def centavos(value) -> int:
    """PHP amount -> integer centavos. Kills the float drift in the source."""
    if value is None or value == "":
        return 0
    if isinstance(value, str):
        value = value.replace("₱", "").replace("P", "").replace(",", "").strip()
        if not value:
            return 0
    try:
        return round(float(value) * 100)
    except (TypeError, ValueError):
        return 0


def text(value) -> str:
    return "" if value is None else str(value).strip()


def iso(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def column(ws, col: int, first: int, last: int) -> list[str]:
    """Non-empty trimmed values from a reference-list column."""
    out = []
    for r in range(first, last + 1):
        v = text(ws.cell(r, col).value)
        if v:
            out.append(v)
    return out


# ── extractors ─────────────────────────────────────────────────────────────

def read_ledger(ws, start_row: int, id_col: int, date_col: int) -> list[dict]:
    """Read a ledger-shaped sheet. Column layout is identical for DATABASE and
    DELETED DATA apart from a leading timestamp column on the latter."""
    rows = []
    r = start_row
    blanks = 0
    while blanks < 5:
        if ws.cell(r, id_col).value is None and ws.cell(r, date_col).value is None:
            blanks += 1
            r += 1
            continue
        blanks = 0
        off = date_col - 3  # offset of this sheet's columns vs DATABASE's

        d = iso(ws.cell(r, date_col).value)
        if d is None:
            r += 1
            continue

        amount = centavos(ws.cell(r, 10 + off).value)
        fee = centavos(ws.cell(r, 11 + off).value)
        total = centavos(ws.cell(r, 12 + off).value)

        rows.append({
            "recordNumber": int(ws.cell(r, id_col).value or 0),
            "date": d,
            "type": text(ws.cell(r, 4 + off).value),
            "fromWallet": text(ws.cell(r, 5 + off).value),
            "toWallet": text(ws.cell(r, 6 + off).value),
            "category": text(ws.cell(r, 7 + off).value),
            "item": text(ws.cell(r, 8 + off).value),
            "description": text(ws.cell(r, 9 + off).value),
            "amount": amount,
            "fee": fee,
            # `total` is stored, but assert the invariant rather than trust it
            "total": total if total else amount + fee,
            "notes": text(ws.cell(r, 13 + off).value),
            "status": text(ws.cell(r, 14 + off).value),
        })
        r += 1
    return rows


def read_reference(ws) -> dict:
    spending_types = []
    for r in range(7, 57):
        name = text(ws.cell(r, 16).value)          # P
        if name:
            spending_types.append({
                "name": name,
                "remark": text(ws.cell(r, 17).value),  # Q
            })
    return {
        "wallets": column(ws, 3, 7, 56),            # C  ACTIVE WALLET
        "savings": column(ws, 4, 7, 56),            # D  SAVINGS
        "bills": column(ws, 8, 7, 56),              # H  BILLS
        "subscriptions": column(ws, 9, 7, 56),      # I  SUBSCRIPTION
        "revenueCategories": column(ws, 13, 7, 56),  # M  CATEGORIES
        "spendingTypes": spending_types,            # P + Q
        # column U (API KEY) is deliberately not read
    }


def read_budgets(ws, year: int) -> dict:
    """BUDGETING H11:S11 spending, H12:S12 bills & subscriptions."""
    return {
        str(year): {
            "spending": [centavos(ws.cell(11, 8 + i).value) for i in range(12)],
            "billsSubs": [centavos(ws.cell(12, 8 + i).value) for i in range(12)],
        }
    }


def read_expected(wb) -> dict:
    """Figures straight from the workbook's own cells — the parity targets."""
    inp, ins, smry, bud, be = (
        wb["INPUT PAGE"], wb["INSIGHTS"], wb["SUMMARY"],
        wb["BUDGETING"], wb["BACKEND"],
    )

    wallets, savings = {}, {}
    for r in range(7, 20):
        name, bal = text(inp.cell(r, 9).value), inp.cell(r, 10).value
        if name and bal is not None:
            wallets[name] = centavos(bal)
        name, bal = text(inp.cell(r, 11).value), inp.cell(r, 12).value
        if name and bal is not None:
            savings[name] = centavos(bal)

    ranking = []
    for r in range(73, 123):                       # BACKEND Q/R ranking tally
        name, amt = text(be.cell(r, 17).value), be.cell(r, 18).value
        if name and amt is not None:
            ranking.append({"name": name, "amount": centavos(amt)})

    # MOST USED WALLET — read the SUMMARY panel the user actually sees
    # (Q13:T23), NOT the BACKEND V/W cache, which holds a stale/different
    # metric (it reports Maya 60,943.22 where the panel shows 46,125.43).
    wallet_usage = []
    for r in range(13, 24):
        name, amt = text(smry.cell(r, 17).value), smry.cell(r, 20).value
        if name and amt is not None:
            wallet_usage.append({"name": name, "amount": centavos(amt)})

    monthly = []
    for i in range(12):                             # BUDGETING B18:E29
        monthly.append({
            "month": MONTHS[i],
            "budget": centavos(bud.cell(18 + i, 3).value),
            "spending": centavos(bud.cell(18 + i, 4).value),
            "remaining": centavos(bud.cell(18 + i, 5).value),
        })

    return {
        # The month the workbook was captured in — INSIGHTS figures are for it.
        "asOf": iso(inp.cell(2, 13).value) or iso(smry.cell(3, 6).value),
        "walletBalances": wallets,
        "savingsBalances": savings,
        "insights": {
            "spendingBudget": centavos(ins.cell(14, 9).value),   # I14
            "billsSubsBudget": centavos(ins.cell(15, 9).value),  # I15
            "totalSpendThisMonth": centavos(ins.cell(16, 9).value),  # I16
            "status": text(ins.cell(17, 9).value),               # I17
        },
        "summary": {
            "spending": centavos(smry.cell(4, 4).value),      # D4
            "revenue": centavos(smry.cell(5, 4).value),       # D5
            "subscription": centavos(smry.cell(6, 4).value),  # D6
            "bills": centavos(smry.cell(7, 4).value),         # D7
            "savings": centavos(smry.cell(8, 4).value),       # D8
            "totalFunds": centavos(smry.cell(9, 4).value),    # D9
        },
        "spendingRanking": ranking,
        "walletUsage": wallet_usage,
        "monthlyBudgetSummary": monthly,
        "forecast": {
            "spending": [centavos(bud.cell(19, 8 + i).value) for i in range(12)],
            "billsSubs": [centavos(bud.cell(20, 8 + i).value) for i in range(12)],
        },
        "netCashFlow": {
            "revenue":  [centavos(bud.cell(27, 8 + i).value) for i in range(12)],
            "expense":  [centavos(bud.cell(28, 8 + i).value) for i in range(12)],
            "savings":  [centavos(bud.cell(29, 8 + i).value) for i in range(12)],
            "transfer": [centavos(bud.cell(30, 8 + i).value) for i in range(12)],
        },
    }


# ── main ───────────────────────────────────────────────────────────────────

def main() -> int:
    if not WORKBOOK.exists():
        sys.exit(f"Workbook not found:\n  {WORKBOOK}")

    print(f"Reading (read-only): {WORKBOOK.name}")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=False)

    transactions = read_ledger(wb["DATABASE"], DATA_START_ROW, 2, 3)
    deleted = read_ledger(wb["DELETED DATA"], DATA_START_ROW, 3, 4)
    reference = read_reference(wb["CATEGORIES"])
    expected = read_expected(wb)

    year = int(transactions[0]["date"][:4]) if transactions else 2026
    budgets = read_budgets(wb["BUDGETING"], year)

    # Invariant from the analysis: total == amount + fee, for every row.
    bad = [t for t in transactions if t["total"] != t["amount"] + t["fee"]]
    if bad:
        print(f"  WARNING: {len(bad)} rows where total != amount + fee:")
        for t in bad[:5]:
            print(f"    #{t['recordNumber']} {t['date']} "
                  f"{t['amount']} + {t['fee']} != {t['total']}")

    fixture = {
        "_meta": {
            "source": WORKBOOK.name,
            "extractedAt": datetime.now().isoformat(timespec="seconds"),
            "note": "All money is INTEGER CENTAVOS. Regenerate with "
                    "`python tools/extract_fixture.py`.",
            "warning": "Contains real financial data. Never commit — see .gitignore.",
        },
        "transactions": transactions,
        "deleted": deleted,
        "budgets": budgets,
        "reference": reference,
        "expected": expected,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(fixture, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"""
  transactions      {len(transactions)}
  deleted           {len(deleted)}
  date range        {transactions[0]['date']} -> {transactions[-1]['date']}
  wallets/savings   {len(reference['wallets'])}/{len(reference['savings'])}
  spending types    {len(reference['spendingTypes'])}
  parity targets    {len(expected['walletBalances'])} wallet balances, \
{len(expected['spendingRanking'])} ranking rows
  invariant         total == amount + fee  {'OK' if not bad else f'{len(bad)} VIOLATIONS'}

  -> {OUT.relative_to(ROOT)}""")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
